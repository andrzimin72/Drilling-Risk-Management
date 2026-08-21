"""
Spreadsheet Parser
Parses Excel workbooks (.xlsx, .xls) for oil and gas engineering data.
Bilingual (English/Russian) sheet classification.
"""
from __future__ import annotations
from pathlib import Path
from typing import Any

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

SHEET_TYPE_SIGNALS: dict[str, list[str]] = {
    "completion_design": ["stage", "cluster", "perforation", "proppant", "frac", "fluid",
                          "стадия", "кластер", "перфорация", "проппант", "грп", "жидкость"],
    "drilling_parameters": ["rop", "wob", "rpm", "depth", "mud weight", "bit", "formation",
                            "механическая скорость", "нагрузка", "обороты", "глубина", "плотность", "долото"],
    "production_data": ["oil", "gas", "water", "rate", "choke", "gor", "bopd",
                        "дебит", "нефть", "газ", "вода", "обводненность", "газовый фактор"],
    "survey_data": ["md", "inc", "az", "tvd", "northing", "easting", "dogleg",
                    "зенитный угол", "азимут", "глубина по стволу", "твг", "интенсивность"],
    "mud_report": ["mud weight", "viscosity", "pv", "yp", "gel", "ph", "chloride",
                   "плотность", "вязкость", "снг", "статическое"],
    "casing_design": ["casing", "od", "weight", "grade", "top", "shoe", "cement",
                      "колонна", "диаметр", "цемент", "башмак"],
    "bha_sheet": ["bha", "component", "od", "id", "length", "serial", "кпбт", "забойная компоновка"],
}

class SpreadsheetParser:
    """Parse Excel workbooks for oil and gas engineering data."""
    def extract_text(self, path: Path) -> str:
        texts = []
        if HAS_OPENPYXL:
            try:
                wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    texts.append(f"=== Sheet: {sheet_name} ===")
                    for row in ws.iter_rows(values_only=True, max_row=500):
                        row_vals = [str(c) for c in row if c is not None and str(c).strip()]
                        if row_vals: texts.append(" | ".join(row_vals))
                wb.close()
            except Exception as exc:
                texts.append(f"[OPENPYXL ERROR: {exc}]")
        elif HAS_PANDAS:
            try:
                xl = pd.ExcelFile(str(path))
                for sheet_name in xl.sheet_names:
                    df = xl.parse(sheet_name, nrows=500)
                    texts.append(f"=== Sheet: {sheet_name} ===")
                    texts.append(df.to_string(index=False, na_rep=""))
            except Exception as exc:
                texts.append(f"[PANDAS ERROR: {exc}]")
        else:
            texts.append("[NO EXCEL PARSER AVAILABLE]")
        return "\n".join(texts)

    def extract_structured(
        self, path: Path, raw_text: str = ""
    ) -> tuple[dict[str, Any], list[dict], list[dict]]:
        if not HAS_OPENPYXL and not HAS_PANDAS:
            return ({"drilling": {}, "directional": {}, "completions": {}, "logs": {}, "production": {}}, [],
                    [{"source_section": "no_excel_parser", "page_or_depth_range": "", "confidence": 0.0}])
                    
        all_sheets = _read_all_sheets(path)
        extracted: dict[str, Any] = {"drilling": {}, "directional": {}, "completions": {}, "logs": {}, "production": {}}
        tables: list[dict[str, Any]] = []
        references: list[dict[str, Any]] = []
        
        for sheet_name, sheet_data in all_sheets.items():
            sheet_type = _classify_sheet(sheet_name, sheet_data)
            table_entry = _sheet_to_table(sheet_name, sheet_data, sheet_type)
            if table_entry: tables.append(table_entry)
            entities = _extract_from_sheet(sheet_data, sheet_type)
            for domain, values in entities.items():
                if domain in extracted: extracted[domain].update(values)
            references.append({"source_section": f"sheet:{sheet_name}", "page_or_depth_range": f"{len(sheet_data)} rows", "confidence": 0.80, "sheet_type": sheet_type})
        return extracted, tables, references

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------
def _read_all_sheets(path: Path) -> dict[str, list[list[Any]]]:
    sheets: dict[str, list[list[Any]]] = {}
    if HAS_OPENPYXL:
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
            for name in wb.sheetnames:
                ws = wb[name]
                rows = [[c for c in row] for row in ws.iter_rows(values_only=True, max_row=2000) if any(c is not None for c in row)]
                sheets[name] = rows
            wb.close()
        except Exception: pass
    elif HAS_PANDAS:
        try:
            xl = pd.ExcelFile(str(path))
            for name in xl.sheet_names:
                df = xl.parse(name, nrows=2000, header=None)
                sheets[name] = df.values.tolist()
        except Exception: pass
    return sheets

def _classify_sheet(name: str, data: list[list[Any]]) -> str:
    name_lower = name.lower()
    sample_text = " ".join(str(cell).lower() for row in data[:5] for cell in row if cell is not None)
    combined = name_lower + " " + sample_text
    best_type, best_score = "unknown", 0
    for sheet_type, signals in SHEET_TYPE_SIGNALS.items():
        score = sum(1 for s in signals if s in combined)
        if score > best_score: best_score, best_type = score, sheet_type
    return best_type if best_score > 0 else "unknown"

def _sheet_to_table(sheet_name: str, data: list[list[Any]], sheet_type: str) -> dict[str, Any] | None:
    if not data: return None
    header_row, data_rows = None, []
    for i, row in enumerate(data):
        if any(cell is not None and str(cell).strip() for cell in row):
            if header_row is None:
                header_row = [str(c).strip() if c is not None else f"col_{j}" for j, c in enumerate(row)]
            else:
                data_rows.append([_clean_cell(c) for c in row])
        if len(data_rows) >= 1000: break
    if not header_row: return None
    return {"name": sheet_name, "type": sheet_type, "columns": header_row, "rows": data_rows[:500], "total_rows": len(data_rows)}

def _clean_cell(value: Any) -> Any:
    if value is None: return None
    if isinstance(value, float) and value != value: return None
    return value

def _extract_from_sheet(data: list[list[Any]], sheet_type: str) -> dict[str, dict[str, Any]]:
    entities: dict[str, dict[str, Any]] = {"drilling": {}, "directional": {}, "completions": {}, "logs": {}, "production": {}}
    if sheet_type == "completion_design":
        flat_vals = {str(row[i]).lower(): _safe_num(row[i + 1]) for row in data for i in range(len(row) - 1) if row[i] is not None and row[i + 1] is not None}
        for key, val in flat_vals.items():
            if val is None: continue
            if "stage" in key or "стадия" in key: entities["completions"]["stage_count"] = val
            if "cluster" in key or "кластер" in key: entities["completions"]["cluster_count"] = val
            if "fluid" in key or "жидкость" in key: entities["completions"]["total_fluid_bbls"] = val
            if "proppant" in key or "проппант" in key: entities["completions"]["total_proppant_lbs"] = val
    if sheet_type == "drilling_parameters":
        for row in data[:50]:
            row_text = " ".join(str(c).lower() for c in row if c is not None)
            if "rop" in row_text or "механическая скорость" in row_text:
                for cell in row:
                    if isinstance(cell, (int, float)) and 0 < cell < 1000:
                        entities["drilling"]["rop_ft_hr" if cell > 50 else "rop_m_hr"] = cell
                        break
            if "mud weight" in row_text or "плотность" in row_text:
                for cell in row:
                    if isinstance(cell, (int, float)) and 1.0 <= cell <= 22.0:
                        entities["drilling"]["mud_weight_ppg" if cell > 6.0 else "mud_weight_sg"] = cell
                        break
    return entities

def _safe_num(val: Any) -> Any:
    try: return float(val)
    except (TypeError, ValueError): return None
